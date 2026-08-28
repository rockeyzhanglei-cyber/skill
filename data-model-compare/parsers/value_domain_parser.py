# -*- coding: utf-8 -*-
"""
值域字典（代码表）解析器

把"值域字典"这一类文档解析成统一结构，供 value_domain_comparator 比对。
与字段结构表（### 表名）不同，值域字典是"代码 -> 值含义"的代码表集合，
本解析器单独处理这一维度，使数据标准比对覆盖"结构 + 值域"两个层面。

统一的值域字典结构：
{
  "<domain_key>": {
     "name": "性别代码表",
     "std_no": "GB/T 2261.1-2003",   # 标准号（可能为空）
     "codes": { "0": "未知的性别", "1": "男性", ... },
     "deletions": [("21","初婚"), ...],  # 文档中明确"删除"的值
     "source_file": "...",
     "raw_heading": "生理性别代码表(GB/T 2261.1-2003)"
  },
  ...
}
"""

import os
import re
import openpyxl


def _norm_std_no(s: str):
    """标准化号归一化：去空格、全角括号转半角，便于跨文档对齐。"""
    if not s:
        return ''
    s = s.strip()
    s = s.replace('（', '(').replace('）', ')')
    s = re.sub(r'\s+', '', s)
    return s


def _norm_name(s: str):
    if not s:
        return ''
    s = s.strip()
    # 去掉末尾的标准号括号，仅保留名称主体
    s = re.sub(r'[（(][^（）()]*[）)]$', '', s).strip()
    return s


def _parse_heading(text: str):
    """从 '名称(标准号)' 形式的标题中拆出 (name, std_no)。"""
    text = text.strip().lstrip('#').strip()
    m = re.match(r'^(.*?)\s*[（(]\s*([^（）()]+?)\s*[）)]\s*$', text)
    if m:
        name = m.group(1).strip()
        std_no = _norm_std_no(m.group(2))
        return name, std_no
    return text, ''


def parse_value_domains_from_md(md_path: str) -> dict:
    """解析目标标准的值域字典 markdown（由 docx 转换而来）。"""
    if not md_path or not os.path.exists(md_path):
        return {}
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.read().split('\n')

    domains = {}
    i = 0
    n = len(lines)
    # 跳过分隔符/目录等前言：从第一个出现的 "## " 代码表标题开始较稳妥，
    # 但为稳健起见仍全量扫描，仅靠"标题后紧跟表格"判断。
    while i < n:
        line = lines[i]
        # 检测标题（#/##/###）
        if not re.match(r'^#{1,4}\s+', line):
            i += 1
            continue
        heading = line.strip()
        # 标题需含"代码"或"表"才视为值域字典标题
        if '代码' not in heading and '表' not in heading:
            i += 1
            continue
        # 向前扫描：跳过空行与"删除XX"等说明行，直到遇到表格或下一个标题
        j = i + 1
        hit_table = False
        while j < n:
            lj = lines[j].strip()
            if lj == '':
                j += 1
                continue
            if lj.startswith('|'):
                hit_table = True
                break
            if re.match(r'^#{1,4}\s+', lj):
                break  # 遇到下一个标题，本标题无表格
            # 其余为说明/删除行，跳过（删除项在下方统一处理）
            j += 1
        if not hit_table or j >= n:
            i += 1
            continue
        name, std_no = _parse_heading(heading)
        # 解析其后的 markdown 表格
        codes = {}
        deletions = []
        # 提取标题与表格之间说明行里的"删除XX-YY"信息
        for di in range(i + 1, j):
            dl = lines[di].strip()
            if '删除' in dl:
                deletions.extend(_extract_deletions(dl))
        table_start = j
        header = [c.strip() for c in lines[table_start].strip().strip('|').split('|')]
        # 表头应含"值/代码"与"值含义/名称"
        if not _looks_like_code_table(header):
            i += 1
            continue
        code_idx, name_idx = _code_name_index(header)
        k = table_start + 1  # 跳过表头
        if k < n and re.match(r'^\s*\|[\s:|-]+\|\s*$', lines[k]):
            k += 1  # 跳过分隔行
        while k < n and lines[k].startswith('|'):
            cells = [c.strip() for c in lines[k].strip().strip('|').split('|')]
            if len(cells) <= max(code_idx, name_idx):
                k += 1
                continue
            code = cells[code_idx].strip()
            meaning = cells[name_idx].strip()
            if code == '' or meaning == '':
                k += 1
                continue
            codes[code] = meaning
            k += 1
        # 将解析到的字典写入（用 name+std_no 作为 key，去重）
        key = _domain_key(name, std_no)
        if key not in domains:
            domains[key] = {
                'name': name,
                'std_no': std_no,
                'codes': codes,
                'deletions': deletions,
                'source_file': os.path.basename(md_path),
                'raw_heading': heading.lstrip('#').strip(),
            }
        i = k
    return domains


def parse_value_domains_from_xlsx(xlsx_path: str) -> dict:
    """解析源标准的值域字典 xlsx（区域平台60 值域字典_V6.0.xxxx.xlsx）。

    主要针对 '数据元值域' 工作表；其余目录类工作表（平台科室目录等）
    作为补充值域一并纳入。
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        return {}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    domains = {}

    target_sheets = [s for s in wb.sheetnames if s == '数据元值域']
    # 若没有数据元值域，则把其余所有表都当作值域来源
    if not target_sheets:
        target_sheets = wb.sheetnames

    for sn in target_sheets:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else '' for c in rows[0]]
        # 找列索引
        name_idx = _col_index(header, ['值域名称', '名称', '字典名称', '代码表'])
        code_sys_idx = _col_index(header, ['值域代码', '标准号', '代码系统', 'code_system'])
        val_code_idx = _col_index(header, ['值代码', '代码', '值', '编码'])
        val_name_idx = _col_index(header, ['值名称', '值含义', '名称', '含义'])
        if val_code_idx is None or val_name_idx is None:
            continue
        for r in rows[1:]:
            if r is None:
                continue
            r = list(r)
            domain_name = r[name_idx].strip() if (name_idx is not None and name_idx < len(r) and r[name_idx]) else ''
            std_no = ''
            if code_sys_idx is not None and code_sys_idx < len(r) and r[code_sys_idx]:
                std_no = _norm_std_no(str(r[code_sys_idx]))
            code = str(r[val_code_idx]).strip() if val_code_idx < len(r) and r[val_code_idx] is not None else ''
            meaning = str(r[val_name_idx]).strip() if val_name_idx < len(r) and r[val_name_idx] is not None else ''
            if not domain_name or code == '' or meaning == '':
                continue
            key = _domain_key(domain_name, std_no)
            d = domains.setdefault(key, {
                'name': domain_name,
                'std_no': std_no,
                'codes': {},
                'deletions': [],
                'source_file': os.path.basename(xlsx_path),
                'raw_heading': f"{domain_name}({std_no})",
                'sheets': set(),
            })
            d['codes'][code] = meaning
            d['sheets'].add(sn)
    # sheets 集合不可序列化，转成列表
    for d in domains.values():
        d['sheets'] = sorted(d.get('sheets', []))
    return domains


def _domain_key(name: str, std_no: str) -> str:
    return f"{std_no}||{name}" if std_no else name


def _looks_like_code_table(header):
    h = ''.join(header)
    return ('值' in h or '代码' in h or '编码' in h) and ('含义' in h or '名称' in h or '值' in h)


def _code_name_index(header):
    code_idx = _col_index(header, ['值', '代码', '编码', 'code', 'Value', 'Code'])
    name_idx = _col_index(header, ['值含义', '含义', '名称', 'name', 'Name', '值名称'])
    if code_idx is None:
        code_idx = 0
    if name_idx is None:
        name_idx = 1 if len(header) > 1 else 0
    return code_idx, name_idx


def _col_index(header, candidates):
    for c in candidates:
        for idx, h in enumerate(header):
            if c == h or c in h:
                return idx
    return None


def _extract_deletions(text: str):
    """从 '删除21-初婚；删除22-再婚' 这类说明中提取删除项 [(code, name), ...]。"""
    out = []
    for m in re.finditer(r'删除\s*([0-9A-Za-z_.．·-]+)\s*[-－]?\s*([^；;，,]*?)(?=；|;|$)', text):
        code = m.group(1).strip()
        name = m.group(2).strip().strip('；;，,').strip()
        if code:
            out.append((code, name))
    return out


def _norm_meaning(s: str):
    """值含义归一化：去空格、去句末'的'、去标点，用于判定名称是否实质相同。"""
    if not s:
        return ''
    s = s.strip()
    s = s.replace('（', '(').replace('）', ')')
    s = re.sub(r'[\s、，,。.;；:：]', '', s)
    s = s.replace('的', '').replace('之', '')
    return s


def parse_value_domains_from_flat_md(md_path: str) -> dict:
    """解析 xlsx 转换而来的「扁平」值域字典 markdown。

    典型结构：以 '# 数据元值域' 为表名，其后紧跟一张宽表，列含
    '类别 | 值域名称 | 值域代码 | 值代码 | 值名称 | ...'。
    与 parse_value_domains_from_md 的 '# 代码表' 标题块不同，这里所有值域
    平铺在同一张表里，因此单独处理。

    返回结构同 parse_value_domains_from_md：
        { "值域代码||值域名称": {"name":..., "std_no":..., "codes": {code: name}} }
    """
    if not md_path or not os.path.exists(md_path):
        return {}
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.read().split('\n')

    domains = {}
    in_table = False
    idx = {}
    for line in lines:
        s = line.strip()
        if s.startswith('#'):
            in_table = (s.lstrip('#').strip() == '数据元值域')
            idx = {}
            continue
        if not in_table or not s.startswith('|'):
            continue
        cells = [c.strip() for c in s.strip('|').split('|')]
        if not idx:
            # 解析表头，定位列
            for i, h in enumerate(cells):
                if h == '值域名称':
                    idx['name'] = i
                elif h == '值域代码':
                    idx['code'] = i
                elif h == '值代码':
                    idx['val'] = i
                elif h == '值名称':
                    idx['mean'] = i
            continue
        if set(s) <= set('|:- '):  # 分隔行
            continue
        name = cells[idx['name']] if 'name' in idx and idx['name'] < len(cells) else ''
        code = cells[idx['code']] if 'code' in idx and idx['code'] < len(cells) else ''
        val = cells[idx['val']] if 'val' in idx and idx['val'] < len(cells) else ''
        mean = cells[idx['mean']] if 'mean' in idx and idx['mean'] < len(cells) else ''
        if not name or not val or not mean:
            continue
        key = f"{code}||{name}" if code else name
        d = domains.setdefault(key, {'name': name, 'std_no': code, 'codes': {}})
        d['codes'][val] = mean
    return domains


def _split_std_no_name(rest: str):
    """从 'GB/T 2261.1-2003 生理性别代码' / '药品编码 STD_MEDICINE' 这类
    小节标题中拆出 (name, std_no)。标准号位置不固定，需双向识别。"""
    rest = rest.strip()
    if not rest:
        return '', ''
    # 候选标准号模式（顺序无关）
    pat = (r'GB/T\s*[\d.\-]+|WS/T?\s*[\d.\-]+|STD_[A-Za-z0-9_.]+|'
           r'CT\s*[\d.]+|[A-Z]{2,}\s*[\d.\-]+')
    m = re.search(pat, rest)
    if m:
        std_no = re.sub(r'\s+', '', m.group(0))  # 去空格，便于跨文档对齐
        name = (rest[:m.start()] + rest[m.end():]).strip()
        # 去掉残留的标准号片段
        name = re.sub(pat, '', name).strip()
        name = name.strip(' 、，,。；;：:·')
        return name, std_no
    return rest, ''


def _xj_header_indices(header):
    """为『代码|类型|备注』类表头定位 代码列 与 含义列。"""
    code_idx = None
    for idx, h in enumerate(header):
        if any(k in h for k in ('代码', '值代码', '编码', '值')):
            code_idx = idx
            break
    if code_idx is None:
        if header:
            code_idx = 0
        else:
            return None, None
    name_idx = None
    for idx, h in enumerate(header):
        if idx == code_idx:
            continue
        if any(k in h for k in ('名称', '含义', '类型', '说明', '意思', '描述')):
            name_idx = idx
            break
    if name_idx is None:
        name_idx = (code_idx + 1) if code_idx + 1 < len(header) else code_idx
    return code_idx, name_idx


def parse_value_domains_from_sectioned_md(md_path: str) -> dict:
    """解析『编号小节标题 + 后续代码表』形式的值域字典 markdown。

    典型来源：新疆维吾尔自治区...值域字典V1.1.5.docx 转换出的 MD。
    结构：
        3.1.1  GB/T 2261.1-2003 生理性别代码
        | 代码 | 类型 | 备注 |
        | 0 | 未知的性别 |  |
        ...
    小节标题为纯文本（非 '#' 标题），以 '数字.数字[.数字]' 开头，
    其后跟『标准号 + 中文域名』（顺序不固定）。
    """
    if not md_path or not os.path.exists(md_path):
        return {}
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.read().split('\n')
    n = len(lines)
    domains = {}
    sec_re = re.compile(r'^\s*(\d+\.\d+(?:\.\d+)*)\s+(.*)$')
    i = 0
    while i < n:
        line = lines[i].rstrip('\n')
        m = sec_re.match(line.strip())
        if not m:
            i += 1
            continue
        rest = m.group(2).strip()
        name, std_no = _split_std_no_name(rest)
        if not name:
            i += 1
            continue
        # 向后找代码表（遇到下一个小节标题或无表则放弃）
        j = i + 1
        table_start = None
        while j < n:
            lj = lines[j].strip()
            if lj == '':
                j += 1
                continue
            if sec_re.match(lj):
                break
            if lj.startswith('|'):
                table_start = j
                break
            j += 1
        if table_start is None or table_start >= n:
            i += 1
            continue
        header = [c.strip() for c in lines[table_start].strip().strip('|').split('|')]
        code_idx, name_idx = _xj_header_indices(header)
        if code_idx is None:
            i += 1
            continue
        k = table_start + 1
        if k < n and re.match(r'^\s*\|[\s:|-]+\|\s*$', lines[k]):
            k += 1
        codes = {}
        while k < n and lines[k].startswith('|'):
            cells = [c.strip() for c in lines[k].strip().strip('|').split('|')]
            if len(cells) <= max(code_idx, name_idx):
                k += 1
                continue
            code = cells[code_idx].strip()
            meaning = cells[name_idx].strip()
            if code == '' or meaning == '':
                k += 1
                continue
            codes[code] = meaning
            k += 1
        if codes:
            key = _domain_key(name, std_no)
            if key not in domains:
                domains[key] = {
                    'name': name,
                    'std_no': std_no,
                    'codes': codes,
                    'deletions': [],
                    'source_file': os.path.basename(md_path),
                    'raw_heading': rest,
                }
        # 跳到当前域表末尾继续
        i = k if k > table_start else (j if j > i else i + 1)
    return domains


def build_value_domain_index(domains: dict) -> dict:
    """从解析结果构建「按名称查找」与「按值域代码查找」两个索引。

    返回 {'by_name': {域名: codes}, 'by_cv': {值域代码: codes}}，
    便于把字段描述里的 'CVxx[域名]' 关联到具体代码表。
    """
    by_name, by_cv = {}, {}
    for key, v in domains.items():
        name = key.split('||', 1)[1] if '||' in key else key
        if v.get('codes'):
            by_name[name] = v['codes']
            if v.get('std_no'):
                by_cv[v['std_no']] = v['codes']
    return {'by_name': by_name, 'by_cv': by_cv}
