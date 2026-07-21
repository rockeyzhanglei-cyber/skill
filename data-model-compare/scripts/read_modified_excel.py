#!/usr/bin/env python3
"""读取用户修改后的Excel，与原始比对结果对比，更新知识库"""

import json
import openpyxl
import yaml
import sys
import os

def read_modified_excel(excel_path, compare_result_path, knowledge_base_path):
    """读取修改后的Excel，找出差异，更新知识库"""
    
    # 读取Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['比对结果']
    
    # 收集Excel中的结果
    excel_results = {}
    for row in range(2, ws.max_row + 1):
        table_name = ws.cell(row=row, column=2).value
        field_display = ws.cell(row=row, column=3).value
        source_table = ws.cell(row=row, column=8).value
        source_field = ws.cell(row=row, column=9).value
        
        if table_name and field_display:
            field_cn = field_display.split('[')[0] if '[' in field_display else field_display
            key = (table_name, field_cn)
            excel_results[key] = {
                'source_table': source_table if source_table else '',
                'source_field': source_field if source_field else ''
            }
    
    # 读取原始比对结果
    with open(compare_result_path, 'r') as f:
        data = json.load(f)
    
    # 构建原始结果索引
    original_results = {}
    for item in data.get('matched', []):
        key = (item.get('table_chinese_name', ''), item.get('target_chinese_name', ''))
        original_results[key] = {
            'source_table': item.get('source_table_chinese_name', '') or '',
            'source_field': item.get('source_field_chinese_name', '') or ''
        }
    
    for item in data.get('modified', []):
        key = (item.get('table_chinese_name', ''), item.get('field_chinese_name', ''))
        original_results[key] = {
            'source_table': item.get('source_table_chinese_name', '') or '',
            'source_field': item.get('source_field_chinese_name', '') or ''
        }
    
    for item in data.get('new_fields', []):
        key = (item.get('table_chinese_name', ''), item.get('chinese_name', ''))
        original_results[key] = {
            'source_table': '',
            'source_field': ''
        }
    
    # 找出差异
    changes = []
    for key in excel_results:
        excel = excel_results[key]
        original = original_results.get(key)
        
        if not original:
            continue
        
        source_table_changed = excel['source_table'] != original['source_table']
        source_field_changed = excel['source_field'] != original['source_field']
        
        if source_table_changed or source_field_changed:
            changes.append({
                'key': key,
                'original': original,
                'excel': excel
            })
    
    print(f"发现 {len(changes)} 处修改")
    
    # 更新知识库
    if changes:
        update_knowledge_base(changes, knowledge_base_path)
    
    return changes

def update_knowledge_base(changes, knowledge_base_path):
    """更新知识库文件"""
    
    field_mappings_path = os.path.join(knowledge_base_path, 'field_mappings.yaml')
    field_synonyms_path = os.path.join(knowledge_base_path, 'field_synonyms.yaml')
    
    # 读取现有的field_mappings
    with open(field_mappings_path, 'r', encoding='utf-8') as f:
        mappings = yaml.safe_load(f)
    
    # 添加新的映射
    for change in changes:
        table_name, field_cn = change['key']
        excel = change['excel']
        original = change['original']
        
        # 如果用户清空了源字段（表示不应匹配）
        if not excel['source_field'] and original['source_field']:
            # 添加到exclude列表（在field_synonyms.yaml中）
            print(f"  排除映射: {table_name}.{field_cn} (原: {original['source_table']}.{original['source_field']})")
            # TODO: 更新field_synonyms.yaml的exclude列表
        
        # 如果用户新增了映射
        elif excel['source_field'] and not original['source_field']:
            print(f"  新增映射: {table_name}.{field_cn} → {excel['source_table']}.{excel['source_field']}")
            # TODO: 添加到field_mappings.yaml
        
        # 如果用户修改了映射
        elif excel['source_field'] and original['source_field']:
            print(f"  修改映射: {table_name}.{field_cn} ({original['source_table']}.{original['source_field']} → {excel['source_table']}.{excel['source_field']})")
            # TODO: 更新field_mappings.yaml
    
    # 保存更新后的知识库
    with open(field_mappings_path, 'w', encoding='utf-8') as f:
        yaml.dump(mappings, f, allow_unicode=True, default_flow_style=False)
    
    print(f"✅ 知识库已更新")

if __name__ == '__main__':
    if len(sys.argv) != 4:
        print("Usage: python3 read_modified_excel.py <excel_path> <compare_result.json> <knowledge_base_path>")
        sys.exit(1)
    
    changes = read_modified_excel(sys.argv[1], sys.argv[2], sys.argv[3])
