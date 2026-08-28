#!/usr/bin/env python3
"""生成可编辑的Excel文件，用于人工核对"""

import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
import re
import sys

def clean_name(name):
    name = re.sub(r'[\*\(\)/\\?\[\]\s]', '', name)
    name = re.sub(r'[—–\-]', '_', name)
    name = re.sub(r'[,，;；]', '_', name)
    name = name.strip()
    if name and name[0].isdigit():
        name = '_' + name
    if not name:
        name = '_empty'
    return name

def generate_excel(compare_result_path, target_standard_path, source_standard_path, output_path):
    with open(compare_result_path, 'r') as f:
        data = json.load(f)
    with open(target_standard_path, 'r') as f:
        target = json.load(f)
    with open(source_standard_path, 'r') as f:
        source = json.load(f)
    
    # 获取全表新增的表名列表（分别存储英文名和中文名）
    new_table_names = set()
    new_table_chinese_names = set()
    for nt in data.get('new_tables', []):
        new_table_names.add(nt.get('table_name', ''))
        new_table_chinese_names.add(nt.get('chinese_name', ''))
    
    # 构建英文名→中文名映射
    table_en_to_cn = {}
    for item in data.get('matched', []):
        en = item.get('source_table', '')
        cn = item.get('source_table_chinese_name', '')
        if en and cn:
            table_en_to_cn[en] = cn
    for item in data.get('modified', []):
        en = item.get('source_table', '')
        cn = item.get('source_table_chinese_name', '')
        if en and cn:
            table_en_to_cn[en] = cn
    
    wb = openpyxl.Workbook()
    
    header_font = Font(name='Times New Roman', size=12, bold=True, color="FFFFFF")
    normal_font = Font(name='Times New Roman', size=12)
    
    header_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws = wb.active
    ws.title = '比对结果'
    
    headers = ["序号", "目标表名", "目标字段", "类型/长度/约束",
               "说明", "值域", "比对结果", "对应源表★", "对应源字段★"]
    col_widths = [6, 28, 30, 16, 50, 12, 22, 28, 28]
    
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.row_dimensions[1].height = 30
    
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    ws.conditional_formatting.add('A2:I10000', 
        FormulaRule(formula=['OR(AND($H2<>"", $I2<>""), AND($G2="满足(字典关联)", $H2=""))'], fill=green_fill))
    ws.conditional_formatting.add('A2:I10000', 
        FormulaRule(formula=['AND($H2<>"", $I2="")'], fill=red_fill))
    
    ws_ref = wb.create_sheet('SourceFields')
    source_tables_info = []
    col = 1
    for table in source.get('tables', []):
        tname = table.get('name', '')  # 修复: table_name -> name
        tcname = table.get('chinese_name', '')
        table_display = tcname if tcname else tname
        ws_ref.cell(row=1, column=col, value=table_display).font = Font(name='Times New Roman', size=12, bold=True)
        fields = table.get('fields', [])
        for i, field in enumerate(fields):
            fcn = field.get('chinese_name', '') or field.get('name', '')  # 修复: field_name -> name
            ws_ref.cell(row=i+2, column=col, value=fcn).font = normal_font
        clean_tname = clean_name(table_display)
        source_tables_info.append({'display_name': table_display, 'clean_name': clean_tname, 'col': col, 'field_count': len(fields)})
        col += 1
    
    for info in source_tables_info:
        col_letter = get_column_letter(info['col'])
        range_str = f"SourceFields!${col_letter}$2:${col_letter}${info['field_count']+1}"
        try:
            dn = DefinedName(info['clean_name'], attr_text=range_str)
            wb.defined_names.add(dn)
        except:
            pass
    
    ws_list = wb.create_sheet('源表列表')
    ws_list.cell(row=1, column=1, value='源表名称').font = Font(name='Times New Roman', size=12, bold=True)
    for i, info in enumerate(source_tables_info, 2):
        ws_list.cell(row=i, column=1, value=info['clean_name']).font = normal_font
    wb.defined_names.add(DefinedName('SourceTableList', attr_text=f"源表列表!$A$2:$A${len(source_tables_info)+1}"))
    
    source_table_validation = DataValidation(type="list", formula1="=SourceTableList", allow_blank=True)
    ws.add_data_validation(source_table_validation)
    
    source_field_validation = DataValidation(type="list", formula1="=INDIRECT(H2)", allow_blank=True)
    ws.add_data_validation(source_field_validation)
    
    match_type_map = {
        'exact_chinese': '满足(精确匹配-中文)', 'exact_english': '满足(精确匹配-英文)',
        'synonym': '满足(同义词匹配)', 'semantic': '满足(语义匹配)', 'keyword': '满足(关键词匹配)',
        'cross_table(1hop)': '满足(跨表匹配)', 'cross_table(2hop)': '满足(跨表匹配)',
        'cross_table(3hop)': '满足(跨表匹配)', 'dictionary': '满足(字典关联)',
        'semantic_mapping': '满足(语义映射)', 'standard_reference': '满足(标准引用)',
        'control_field': '满足(控制字段映射)', 'numbered_field_group': '满足(编号字段组)'
    }
    
    match_type_modified_map = {
        'exact_chinese': '修改(精确匹配-中文)', 'synonym': '修改(同义词匹配)',
        'semantic': '修改(语义匹配)', 'keyword': '修改(关键词匹配)',
        'cross_table(1hop)': '修改(跨表匹配)', 'cross_table(2hop)': '修改(跨表匹配)',
        'cross_table(3hop)': '修改(跨表匹配)', 'semantic_mapping': '修改(语义映射)'
    }
    
    lookup = {}
    for item in data.get('matched', []):
        table_name = item.get('table_chinese_name', '') or item.get('table_name', '')
        field_cn = item.get('target_chinese_name', '')
        key = (table_name, field_cn)
        mt = match_type_map.get(item.get('match_type', ''), '满足(精确匹配-中文)')
        source_table = item.get('source_table_chinese_name', '')
        source_field = item.get('source_field_chinese_name', '')
        cond = item.get('condition_display', '')
        if cond:
            source_field = f"{source_field}（{cond}）"
        lookup[key] = (mt, source_table, source_field, False)
    
    for item in data.get('modified', []):
        table_name = item.get('table_chinese_name', '') or item.get('table_name', '')
        field_cn = item.get('field_chinese_name', '')
        key = (table_name, field_cn)
        mt = match_type_modified_map.get(item.get('match_type', ''), '修改(精确匹配-中文)')
        source_table = item.get('source_table_chinese_name', '')
        source_field = item.get('source_field_chinese_name', '')
        cond = item.get('condition_display', '')
        if cond:
            source_field = f"{source_field}（{cond}）"
        lookup[key] = (mt, source_table, source_field, False)
    
    for item in data.get('new_fields', []):
        table_name = item.get('table_chinese_name', '') or item.get('table_name', '')
        field_cn = item.get('chinese_name', '')
        key = (table_name, field_cn)

        # 检查是否是全表新增（同时检查英文名和中文名）
        original_table_name = item.get('table_name', '')
        original_table_cn = item.get('table_chinese_name', '')
        is_new_table = original_table_name in new_table_names or original_table_cn in new_table_chinese_names

        source_table_en = item.get('source_table_name', '')
        source_table_cn = table_en_to_cn.get(source_table_en, source_table_en)

        if is_new_table:
            lookup[key] = ('新增', '', '', True)
        else:
            lookup[key] = ('新增', source_table_cn, '', False)
    
    def get_field_info(standard, table_name, field_name, field_cn):
        for t in standard.get('tables', []):
            if t.get('chinese_name', '') == table_name or t.get('name', '') == table_name:  # 修复: table_name -> name
                for f in t.get('fields', []):
                    if f.get('name', '') == field_name or f.get('chinese_name', '') == field_cn:  # 修复: field_name -> name
                        dtype = f.get('data_type', '')
                        length = f.get('length', '')
                        constraint = f.get('constraint', '')
                        type_len_con = f"{dtype}/{length}/{constraint}" if dtype else ""
                        desc = f.get('description', '')[:200] if f.get('description') else ''
                        # 修复: value_domain -> value_domains (复数，且是列表)
                        value_domains = f.get('value_domains', [])
                        vd = '; '.join([f"{v.get('code', '')}:{v.get('name', '')}" for v in value_domains]) if value_domains else ''
                        return type_len_con, desc, vd
        return '', '', ''
    
    row = 2
    seq = 1

    for table in target.get('tables', []):
        table_chinese_name = table.get('chinese_name', '')
        table_en_name = table.get('name', '')  # 修复: table_name -> name
        table_name = table_chinese_name or table_en_name
        fields = table.get('fields', [])

        for field in fields:
            field_name = field.get('name', '')  # 修复: field_name -> name
            field_cn = field.get('chinese_name', '')

            # 尝试用中文名查找，如果找不到再用英文名查找
            key = (table_name, field_cn)
            if key in lookup:
                status, source_table, source_field, is_new_table = lookup[key]
            elif table_en_name:
                key_en = (table_en_name, field_cn)
                if key_en in lookup:
                    status, source_table, source_field, is_new_table = lookup[key_en]
                else:
                    # 检查是否是全表新增的表
                    is_new_table = table_chinese_name in new_table_chinese_names or table_en_name in new_table_names
                    status, source_table, source_field = '新增', '', ''
            else:
                # 检查是否是全表新增的表
                is_new_table = table_chinese_name in new_table_chinese_names
                status, source_table, source_field = '新增', '', ''
            
            data_type = field.get('data_type', '')
            length = field.get('length', '')
            constraint = field.get('constraint', '')
            format_str = field.get('format', '')
            # 优先使用完整的格式信息，如果没有则用 数据类型/长度
            if format_str:
                type_len_con = f"{data_type}/{format_str}/{constraint}" if data_type else ""
            else:
                type_len_con = f"{data_type}/{length}/{constraint}" if data_type else ""
            desc = field.get('description', '')[:200] if field.get('description') else ''
            # 修复: value_domain -> value_domains (复数，且是列表)
            value_domains = field.get('value_domains', [])
            vd = '; '.join([f"{v.get('code', '')}:{v.get('name', '')}" for v in value_domains]) if value_domains else ''
            
            field_display = f"{field_cn}[{field_name}]" if field_name else field_cn
            
            if is_new_table:
                fmt = red_fill
            else:
                fmt = None
            
            ws.cell(row=row, column=1, value=seq).font = normal_font
            if fmt: ws.cell(row=row, column=1).fill = fmt
            ws.cell(row=row, column=2, value=table_name).font = normal_font
            if fmt: ws.cell(row=row, column=2).fill = fmt
            ws.cell(row=row, column=3, value=field_display).font = normal_font
            if fmt: ws.cell(row=row, column=3).fill = fmt
            ws.cell(row=row, column=4, value=type_len_con).font = normal_font
            if fmt: ws.cell(row=row, column=4).fill = fmt
            ws.cell(row=row, column=5, value=desc).font = normal_font
            if fmt: ws.cell(row=row, column=5).fill = fmt
            ws.cell(row=row, column=6, value=vd).font = normal_font
            if fmt: ws.cell(row=row, column=6).fill = fmt
            ws.cell(row=row, column=7, value=status).font = normal_font
            if fmt: ws.cell(row=row, column=7).fill = fmt
            ws.cell(row=row, column=8, value=source_table).font = normal_font
            if fmt: ws.cell(row=row, column=8).fill = fmt
            ws.cell(row=row, column=9, value=source_field).font = normal_font
            if fmt: ws.cell(row=row, column=9).fill = fmt
            
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = center_align
            
            desc_chars = len(desc) if desc else 0
            chars_per_line = 25
            lines = max(1, (desc_chars + chars_per_line - 1) // chars_per_line)
            row_height = max(25, lines * 18)
            ws.row_dimensions[row].height = row_height
            
            row += 1
            seq += 1
    
    last_row = row - 1
    source_table_validation.add(f"H2:H{last_row}")
    source_field_validation.add(f"I2:I{last_row}")
    
    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✅ Excel已生成: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) != 5:
        print("Usage: python3 generate_excel.py <compare_result.json> <target_standard.json> <source_standard.json> <output.xlsx>")
        sys.exit(1)
    
    generate_excel(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
