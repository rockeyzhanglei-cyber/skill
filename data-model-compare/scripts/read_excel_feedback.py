#!/usr/bin/env python3
"""
Excel 回写处理器

读取用户编辑后的 Excel 文件，检测映射关系变更，更新知识库。

核心流程：
1. 读取用户编辑后的 Excel（比对结果 sheet）
2. 读取原始 compare_result.json（比对程序的输出）
3. 逐行对比，检测用户的修改
4. 分类变更：修改映射 / 新增字段 / 整表新增 / 取消映射
5. 更新 user_custom_mappings.yaml（本次任务生效）
6. 分析可沉淀到公共知识库的映射（供后续任务复用）

用法：
    python3 read_excel_feedback.py \
        --excel <编辑后的Excel路径> \
        --compare-result <compare_result.json> \
        --target-standard <target_standard.json> \
        --source-standard <source_standard.json> \
        --task-dir <任务目录>
"""

import json
import os
import re
import sys
import argparse
from typing import Dict, List, Tuple, Optional

import yaml
import openpyxl


# ============================================================================
# 工具函数
# ============================================================================

def clean_name(name: str) -> str:
    """清理名称用于 DefinedName（与 generate_excel.py 保持一致）"""
    if not name:
        return '_empty'
    name = re.sub(r'[\*\(\)/\\?\[\]\s]', '', name)
    name = re.sub(r'[—–\-]', '_', name)
    name = re.sub(r'[，,;；]', '_', name)
    name = name.strip()
    if name and name[0].isdigit():
        name = '_' + name
    return name or '_empty'


def parse_field_display(display: str) -> Tuple[str, str]:
    """解析字段显示名 '中文名[英文名]' -> (中文名, 英文名)"""
    if not display:
        return '', ''
    display = str(display).strip()
    match = re.match(r'^(.+?)\[(.+?)\]$', display)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return display, ''


# ============================================================================
# 数据加载
# ============================================================================

def load_json(path: str) -> dict:
    """加载 JSON 文件"""
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_yaml_safe(path: str) -> dict:
    """安全加载 YAML 文件"""
    if not os.path.exists(path):
        return {}
    with open(path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f) or {}


def build_source_table_map(source_standard: dict) -> Dict[str, str]:
    """
    构建源标准表名映射：中文名/英文名 -> 原始英文名

    返回: {display_name: original_name}
    """
    mapping = {}
    for table in source_standard.get('tables', []):
        en_name = table.get('name', '')
        cn_name = table.get('chinese_name', '')
        if en_name:
            mapping[en_name] = en_name
        if cn_name:
            mapping[cn_name] = en_name
            # 也加入 cleaned 版本
            mapping[clean_name(cn_name)] = en_name
    return mapping


def build_source_field_map(source_standard: dict) -> Dict[str, Dict[str, str]]:
    """
    构建源标准字段映射：表英文名 -> {字段中文名/英文名 -> 原始英文名}

    返回: {table_en_name: {field_display: field_en_name}}
    """
    mapping = {}
    for table in source_standard.get('tables', []):
        table_en = table.get('name', '')
        field_map = {}
        for field in table.get('fields', []):
            f_en = field.get('name', '')
            f_cn = field.get('chinese_name', '')
            if f_en:
                field_map[f_en] = f_en
            if f_cn:
                field_map[f_cn] = f_en
        mapping[table_en] = field_map
    return mapping


def build_original_result_index(compare_result: dict) -> Dict[Tuple[str, str], dict]:
    """
    构建原始比对结果索引：(目标表中文名, 目标字段中文名) -> 原始映射信息

    返回: {(table_cn, field_cn): {source_table_cn, source_field_cn, source_table_en, match_type, ...}}
    """
    index = {}

    for item in compare_result.get('matched', []):
        table_cn = item.get('table_chinese_name', '') or item.get('table_name', '')
        field_cn = item.get('target_chinese_name', '')
        key = (table_cn, field_cn)
        index[key] = {
            'source_table_cn': item.get('source_table_chinese_name', '') or '',
            'source_field_cn': item.get('source_field_chinese_name', '') or '',
            'source_table_en': item.get('source_table', '') or '',
            'source_field_en': item.get('source_field', '') or '',
            'match_type': item.get('match_type', ''),
            'original_status': 'matched',
        }

    for item in compare_result.get('modified', []):
        table_cn = item.get('table_chinese_name', '') or item.get('table_name', '')
        field_cn = item.get('field_chinese_name', '')
        key = (table_cn, field_cn)
        index[key] = {
            'source_table_cn': item.get('source_table_chinese_name', '') or '',
            'source_field_cn': item.get('source_field_chinese_name', '') or '',
            'source_table_en': item.get('source_table', '') or '',
            'source_field_en': item.get('source_field', '') or '',
            'match_type': item.get('match_type', ''),
            'original_status': 'modified',
        }

    for item in compare_result.get('new_fields', []):
        table_cn = item.get('table_chinese_name', '') or item.get('table_name', '')
        field_cn = item.get('chinese_name', '')
        key = (table_cn, field_cn)
        index[key] = {
            'source_table_cn': '',
            'source_field_cn': '',
            'source_table_en': '',
            'source_field_en': '',
            'match_type': 'new_field',
            'original_status': 'new_field',
        }

    return index


def get_new_tables_set(compare_result: dict) -> set:
    """获取全表新增的表名集合（中文名和英文名）"""
    result = set()
    for nt in compare_result.get('new_tables', []):
        if nt.get('table_name'):
            result.add(nt['table_name'])
        if nt.get('chinese_name'):
            result.add(nt['chinese_name'])
    return result


# ============================================================================
# Excel 解析
# ============================================================================

def parse_excel_rows(excel_path: str) -> List[dict]:
    """
    解析 Excel 的 '比对结果' sheet

    返回: [{
        'row_num': int,
        'table_name': str,        # 目标表名（中文名或英文名）
        'field_cn': str,          # 目标字段中文名
        'field_en': str,          # 目标字段英文名
        'source_table_raw': str,  # 用户填写的源表（原始值）
        'source_field_raw': str,  # 用户填写的源字段（原始值）
    }]
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 找到 '比对结果' sheet
    ws = None
    for name in wb.sheetnames:
        if '比对结果' in name:
            ws = wb[name]
            break
    if ws is None:
        print(f"  错误: Excel 中没有找到 '比对结果' sheet")
        return []

    rows = []
    for row_num in range(2, ws.max_row + 1):
        # 跳过空行
        table_name = ws.cell(row=row_num, column=2).value
        if not table_name:
            continue

        table_name = str(table_name).strip()
        field_display = str(ws.cell(row=row_num, column=3).value or '').strip()
        source_table_raw = str(ws.cell(row=row_num, column=8).value or '').strip()
        source_field_raw = str(ws.cell(row=row_num, column=9).value or '').strip()

        # None 字符串清理
        if source_table_raw.lower() == 'none':
            source_table_raw = ''
        if source_field_raw.lower() == 'none':
            source_field_raw = ''

        field_cn, field_en = parse_field_display(field_display)

        rows.append({
            'row_num': row_num,
            'table_name': table_name,
            'field_cn': field_cn,
            'field_en': field_en,
            'source_table_raw': source_table_raw,
            'source_field_raw': source_field_raw,
        })

    wb.close()
    return rows


# ============================================================================
# 变更检测
# ============================================================================

def resolve_source_table(raw_value: str, source_table_map: Dict[str, str]) -> Tuple[str, str]:
    """
    将用户填写的源表名解析为 (中文名, 英文名)

    用户可能填写中文名、英文名、或 cleaned 名。
    """
    if not raw_value:
        return '', ''

    # 精确匹配
    if raw_value in source_table_map:
        en_name = source_table_map[raw_value]
        # 反查中文名
        cn_name = ''
        for k, v in source_table_map.items():
            if v == en_name and not k.isascii():
                cn_name = k
                break
        return cn_name, en_name

    # 模糊匹配：用户填写的值包含在某个表名中
    for display, en_name in source_table_map.items():
        if raw_value in display or display in raw_value:
            cn_name = ''
            for k, v in source_table_map.items():
                if v == en_name and not k.isascii():
                    cn_name = k
                    break
            return cn_name, en_name

    # 无法解析，原样返回
    return raw_value, raw_value


def resolve_source_field(raw_value: str, table_en: str,
                         source_field_map: Dict[str, Dict[str, str]]) -> Tuple[str, str]:
    """
    将用户填写的源字段名解析为 (中文名, 英文名)
    """
    if not raw_value:
        return '', ''

    field_map = source_field_map.get(table_en, {})

    # 精确匹配
    if raw_value in field_map:
        en_name = field_map[raw_value]
        cn_name = ''
        for k, v in field_map.items():
            if v == en_name and not k.isascii():
                cn_name = k
                break
        return cn_name, en_name

    # 模糊匹配
    for display, en_name in field_map.items():
        if raw_value in display or display in raw_value:
            cn_name = ''
            for k, v in field_map.items():
                if v == en_name and not k.isascii():
                    cn_name = k
                    break
            return cn_name, en_name

    return raw_value, raw_value


def detect_changes(excel_rows: List[dict],
                   original_index: Dict[Tuple[str, str], dict],
                   source_table_map: Dict[str, str],
                   source_field_map: Dict[str, Dict[str, str]],
                   target_standard: dict,
                   new_tables_set: set) -> List[dict]:
    """
    检测用户编辑与原始结果的差异

    返回: [{
        'type': 'mapping_changed' | 'new_field_assigned' | 'whole_table_new' | 'mapping_cleared',
        'table_cn': str,
        'table_en': str,
        'field_cn': str,
        'field_en': str,
        'original': {...},
        'new_source_table_cn': str,
        'new_source_table_en': str,
        'new_source_field_cn': str,
        'new_source_field_en': str,
    }]
    """
    # 构建目标表映射（用于查找 table_en）
    target_table_map = {}  # display_name -> (cn_name, en_name)
    for table in target_standard.get('tables', []):
        en = table.get('name', '')
        cn = table.get('chinese_name', '')
        if cn:
            target_table_map[cn] = (cn, en)
        if en:
            target_table_map[en] = (cn, en)

    # 按表分组，用于检测"整表新增"
    table_rows = {}  # table_cn -> [rows]
    for row in excel_rows:
        tn = row['table_name']
        # 解析目标表名
        if tn in target_table_map:
            cn, en = target_table_map[tn]
        else:
            cn, en = tn, tn
        table_rows.setdefault(cn, []).append({**row, '_table_cn': cn, '_table_en': en})

    changes = []

    for table_cn, rows in table_rows.items():
        table_en = rows[0]['_table_en'] if rows else ''

        # 检测"整表新增"：该表所有行的源表和源字段都为空
        all_empty = all(
            not r['source_table_raw'] and not r['source_field_raw']
            for r in rows
        )
        if all_empty and len(rows) > 0:
            # 检查原始结果中该表的字段状态
            # original_index 的 key 是 (table_name, field_cn)，需要正确匹配表名
            table_original_statuses = [
                orig.get('original_status', '')
                for (t, f), orig in original_index.items()
                if t == table_cn or t == table_en
            ]

            has_original_match = any(
                status in ('matched', 'modified')
                for status in table_original_statuses
            )
            all_original_new = (
                len(table_original_statuses) > 0 and
                all(status == 'new_field' for status in table_original_statuses)
            )

            # 两种情况都应识别为"整表新增"：
            # 1. 原始结果中有匹配但用户全部清空（用户撤销了匹配）
            # 2. 原始结果中全部是新增，用户清空确认（用户确认整表新增）
            if has_original_match or all_original_new:
                changes.append({
                    'type': 'whole_table_new',
                    'table_cn': table_cn,
                    'table_en': table_en,
                    'field_cn': '',
                    'field_en': '',
                    'original': {},
                    'new_source_table_cn': '',
                    'new_source_table_en': '',
                    'new_source_field_cn': '',
                    'new_source_field_en': '',
                })
                continue

        # 逐行检测
        for row in rows:
            field_cn = row['field_cn']
            field_en = row['field_en']
            source_table_raw = row['source_table_raw']
            source_field_raw = row['source_field_raw']

            # 查找原始结果
            orig = original_index.get((table_cn, field_cn))
            if not orig:
                orig = original_index.get((table_en, field_cn))
            if not orig:
                # 尝试用英文名匹配
                for (t, f), o in original_index.items():
                    if f == field_cn and (t == table_cn or t == table_en):
                        orig = o
                        break
            if not orig:
                continue

            # 解析用户填写的源表和源字段
            new_src_table_cn, new_src_table_en = resolve_source_table(
                source_table_raw, source_table_map
            )
            new_src_field_cn, new_src_field_en = resolve_source_field(
                source_field_raw, new_src_table_en, source_field_map
            )

            # 对比原始值
            orig_src_table_cn = orig.get('source_table_cn', '')
            orig_src_field_cn = orig.get('source_field_cn', '')
            orig_status = orig.get('original_status', '')

            # 判断是否有变更
            table_changed = (source_table_raw != orig_src_table_cn and
                           source_table_raw != orig.get('source_table_en', ''))
            field_changed = (source_field_raw != orig_src_field_cn and
                           source_field_raw != orig.get('source_field_en', ''))

            # 处理空值比较的特殊情况
            if not source_table_raw and not orig_src_table_cn:
                table_changed = False
            if not source_field_raw and not orig_src_field_cn:
                field_changed = False

            if not table_changed and not field_changed:
                continue  # 没有变更

            # 分类变更
            if source_table_raw and not source_field_raw and not orig_src_field_cn:
                # 用户填了源表但没填源字段，且原来也没有 -> 新增字段到指定表
                change_type = 'new_field_assigned'
            elif not source_table_raw and not source_field_raw and orig_src_field_cn:
                # 用户清空了映射 -> 取消映射
                change_type = 'mapping_cleared'
            else:
                # 修改了映射
                change_type = 'mapping_changed'

            changes.append({
                'type': change_type,
                'table_cn': table_cn,
                'table_en': table_en,
                'field_cn': field_cn,
                'field_en': field_en,
                'original': orig,
                'new_source_table_cn': new_src_table_cn,
                'new_source_table_en': new_src_table_en,
                'new_source_field_cn': new_src_field_cn,
                'new_source_field_en': new_src_field_en,
            })

    return changes


# ============================================================================
# 知识库更新
# ============================================================================

def update_user_custom_mappings(changes: List[dict], task_dir: str):
    """
    更新 user_custom_mappings.yaml

    保留原有有效条目，添加/更新用户编辑的条目。
    """
    kb_path = os.path.join(task_dir, 'knowledge_base')
    yaml_path = os.path.join(kb_path, 'user_custom_mappings.yaml')

    # 加载现有文件
    existing = load_yaml_safe(yaml_path)
    existing_mappings = existing.get('mappings', [])

    # 构建索引：(target_table, target_field) -> mapping
    mapping_index = {}
    for m in existing_mappings:
        key = (m.get('target_table', ''), m.get('target_field', ''))
        mapping_index[key] = m

    # 应用变更
    for change in changes:
        table_cn = change['table_cn']
        field_cn = change['field_cn']

        if change['type'] == 'whole_table_new':
            # 整表新增：将该表所有字段标记为新增
            # 这里不逐字段写入，而是在 new_tables 中标记
            existing.setdefault('new_tables', [])
            # 避免重复
            existing_tables = {t.get('table_name', '') for t in existing['new_tables']}
            if table_cn not in existing_tables:
                existing['new_tables'].append({
                    'table_name': table_cn,
                    'reason': '用户确认整表新增',
                })
            continue

        # 构建映射条目
        key = (table_cn, field_cn)
        mapping_entry = {
            'target_table': table_cn,
            'target_field': field_cn,
            'source_table': change['new_source_table_cn'] or change['new_source_table_en'],
            'source_field': change['new_source_field_cn'] or change['new_source_field_en'],
            'source_field_en': change['new_source_field_en'],
            'source_field_cn': change['new_source_field_cn'],
            'original_source_table': change['original'].get('source_table_cn', ''),
            'original_source_field': change['original'].get('source_field_cn', ''),
        }
        mapping_index[key] = mapping_entry

    # 写回文件
    existing['mappings'] = list(mapping_index.values())
    existing.setdefault('description', '用户自定义字段映射（优先级最高）')
    existing.setdefault('version', '2.0')

    os.makedirs(kb_path, exist_ok=True)
    with open(yaml_path, 'w', encoding='utf-8') as f:
        yaml.dump(existing, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    print(f"  已更新 user_custom_mappings.yaml: {len(mapping_index)} 条映射")


def analyze_generalizable(changes: List[dict], skill_dir: str):
    """
    分析可沉淀到公共知识库的映射

    将用户修正的映射分类：
    - 同义词修正：用户修正了 synonym/keyword 匹配错误 → 建议添加到 field_synonyms.yaml
    - 表映射修正：用户修改了表对应关系 → 建议添加到 table_synonyms.yaml
    - 字段映射修正：用户指定了 field_mappings → 建议添加到 field_mappings.yaml
    """
    if not changes:
        return

    suggestions = {
        'synonym_add': [],      # 需要添加的同义词
        'synonym_remove': [],   # 需要移除的同义词
        'table_synonym': [],    # 需要添加的表同义词
        'field_mapping': [],    # 需要添加的字段映射
    }

    for change in changes:
        if change['type'] == 'whole_table_new':
            continue

        orig = change.get('original', {})
        orig_match_type = orig.get('match_type', '')

        if change['type'] == 'mapping_cleared':
            # 用户取消了映射 — 可能是同义词错误
            if orig_match_type in ('synonym', 'keyword', 'semantic'):
                suggestions['synonym_remove'].append({
                    'target_field': change['field_cn'],
                    'wrongly_matched_to': orig.get('source_field_cn', ''),
                    'match_type': orig_match_type,
                })

        elif change['type'] == 'mapping_changed':
            # 用户修改了映射
            if orig_match_type in ('synonym', 'keyword', 'semantic'):
                # 程序匹配错了，用户纠正了 — 可以提取为新的映射规则
                suggestions['field_mapping'].append({
                    'target_table': change['table_cn'],
                    'target_field': change['field_cn'],
                    'source_table': change['new_source_table_cn'],
                    'source_field': change['new_source_field_cn'],
                    'original_match_type': orig_match_type,
                    'correct_source': change['new_source_field_en'],
                })

        elif change['type'] == 'new_field_assigned':
            # 用户为新字段指定了源表 — 可能揭示表映射缺失
            pass

    # 输出建议
    if any(suggestions.values()):
        print("\n  === 知识沉淀建议 ===")
        if suggestions['synonym_remove']:
            print(f"  同义词修正 ({len(suggestions['synonym_remove'])} 条):")
            for s in suggestions['synonym_remove'][:5]:
                print(f"    - \"{s['target_field']}\" 不应匹配 \"{s['wrongly_matched_to']}\" "
                      f"(原匹配方式: {s['match_type']})")
            if len(suggestions['synonym_remove']) > 5:
                print(f"    ... 共 {len(suggestions['synonym_remove'])} 条")

        if suggestions['field_mapping']:
            print(f"  字段映射修正 ({len(suggestions['field_mapping'])} 条):")
            for s in suggestions['field_mapping'][:5]:
                print(f"    - {s['target_table']}.{s['target_field']} -> "
                      f"{s['source_table']}.{s['source_field']} "
                      f"(原匹配方式: {s['original_match_type']})")
            if len(suggestions['field_mapping']) > 5:
                print(f"    ... 共 {len(suggestions['field_mapping'])} 条")

        # 保存到文件供后续处理
        suggestion_path = os.path.join(skill_dir, 'knowledge_base', 'pending_suggestions.yaml')
        with open(suggestion_path, 'w', encoding='utf-8') as f:
            yaml.dump(suggestions, f, allow_unicode=True, default_flow_style=False)
        print(f"\n  建议已保存到: {suggestion_path}")
        print(f"  （这些建议可在下次使用 /skill-creator 时审查和应用）")


# ============================================================================
# 主流程
# ============================================================================

def process_feedback(excel_path: str,
                     compare_result_path: str,
                     target_standard_path: str,
                     source_standard_path: str,
                     task_dir: str,
                     skill_dir: str):
    """处理 Excel 反馈的主入口"""

    print("=" * 60)
    print("Excel 回写处理")
    print("=" * 60)

    # 1. 加载数据
    print("\n[1/5] 加载数据...")
    compare_result = load_json(compare_result_path)
    target_standard = load_json(target_standard_path)
    source_standard = load_json(source_standard_path)

    source_table_map = build_source_table_map(source_standard)
    source_field_map = build_source_field_map(source_standard)
    original_index = build_original_result_index(compare_result)
    new_tables_set = get_new_tables_set(compare_result)

    print(f"  源标准: {len(source_standard.get('tables', []))} 张表")
    print(f"  目标标准: {len(target_standard.get('tables', []))} 张表")
    print(f"  原始比对结果: "
          f"{len(compare_result.get('matched', []))} 满足, "
          f"{len(compare_result.get('modified', []))} 修改, "
          f"{len(compare_result.get('new_fields', []))} 新增")

    # 2. 解析 Excel
    print("\n[2/5] 解析 Excel...")
    excel_rows = parse_excel_rows(excel_path)
    print(f"  读取 {len(excel_rows)} 行数据")

    if not excel_rows:
        print("  错误: Excel 中没有有效数据")
        return []

    # 3. 检测变更
    print("\n[3/5] 检测变更...")
    changes = detect_changes(
        excel_rows, original_index,
        source_table_map, source_field_map,
        target_standard, new_tables_set
    )

    # 统计变更类型
    type_counts = {}
    for c in changes:
        t = c['type']
        type_counts[t] = type_counts.get(t, 0) + 1

    print(f"  检测到 {len(changes)} 处变更:")
    type_labels = {
        'mapping_changed': '修改映射',
        'new_field_assigned': '新增字段指定源表',
        'whole_table_new': '整表新增',
        'mapping_cleared': '取消映射',
    }
    for t, count in type_counts.items():
        label = type_labels.get(t, t)
        print(f"    - {label}: {count} 处")

    if not changes:
        print("\n  没有检测到变更，无需更新知识库。")
        return []

    # 4. 更新知识库
    print("\n[4/5] 更新知识库...")
    update_user_custom_mappings(changes, skill_dir)

    # 5. 分析知识沉淀
    print("\n[5/5] 分析知识沉淀...")
    analyze_generalizable(changes, skill_dir)

    # 输出汇总
    print("\n" + "=" * 60)
    print("处理完成")
    print("=" * 60)
    print(f"\n后续步骤:")
    print(f"  1. 重新运行比对（用户映射将自动生效）:")
    print(f"     python3 {skill_dir}/main.py \\")
    print(f"       --source <原标准> --target <目标标准>")
    print(f"  2. 查看 pending_suggestions.yaml 中的知识沉淀建议")
    print(f"  3. 使用 /skill-creator 审查并应用建议到公共知识库")

    return changes


# ============================================================================
# CLI 入口
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description='Excel 回写处理器')
    parser.add_argument('--excel', '-e', required=True, help='编辑后的 Excel 文件路径')
    parser.add_argument('--compare-result', '-c', required=True, help='compare_result.json 路径')
    parser.add_argument('--target-standard', '-t', required=True, help='target_standard.json 路径')
    parser.add_argument('--source-standard', '-s', required=True, help='source_standard.json 路径')
    parser.add_argument('--task-dir', help='任务目录（默认为 compare_result 所在目录的上级）')
    parser.add_argument('--skill-dir', help='Skill 目录（默认为脚本所在目录的上级）')

    args = parser.parse_args()

    # 推断默认路径
    skill_dir = args.skill_dir or os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    compare_dir = os.path.dirname(os.path.abspath(args.compare_result))
    task_dir = args.task_dir or os.path.dirname(compare_dir)

    # 检查文件存在
    for path, name in [
        (args.excel, 'Excel 文件'),
        (args.compare_result, 'compare_result.json'),
        (args.target_standard, 'target_standard.json'),
        (args.source_standard, 'source_standard.json'),
    ]:
        if not os.path.exists(path):
            print(f"错误: {name} 不存在: {path}")
            sys.exit(1)

    process_feedback(
        excel_path=args.excel,
        compare_result_path=args.compare_result,
        target_standard_path=args.target_standard,
        source_standard_path=args.source_standard,
        task_dir=task_dir,
        skill_dir=skill_dir,
    )


if __name__ == '__main__':
    main()
